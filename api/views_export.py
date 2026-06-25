from django.http import HttpResponse
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
# pyrefly: ignore [missing-import]
from .models import Proyek, Pekerjaan, Aktivitas
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
from reportlab.graphics.shapes import Drawing, Rect
import datetime
import traceback
import logging

logger = logging.getLogger(__name__)


def hitung_progres_proyek(proyek):
    """Hitung progres proyek secara dinamis berdasarkan jumlah aktivitas yang selesai"""
    daftar_aktivitas = Aktivitas.objects.filter(pekerjaan__proyek=proyek)
    total = daftar_aktivitas.count()
    if total == 0:
        return 0
    selesai = daftar_aktivitas.filter(selesai=True).count()
    return round((selesai / total) * 100)


# --- Excel Styles ---
EXCEL_HEADER_FONT = Font(bold=True, color="FFFFFF")
EXCEL_HEADER_FILL = PatternFill("solid", fgColor="1E3A8A")
EXCEL_BOLD_FONT = Font(bold=True)
EXCEL_CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
EXCEL_LEFT_ALIGN = Alignment(horizontal="left", vertical="center")
EXCEL_BORDER = Border(left=Side(style='thin', color="DDDDDD"),
                      right=Side(style='thin', color="DDDDDD"),
                      top=Side(style='thin', color="DDDDDD"),
                      bottom=Side(style='thin', color="DDDDDD"))

def apply_excel_header_style(cell):
    cell.font = EXCEL_HEADER_FONT
    cell.fill = EXCEL_HEADER_FILL
    cell.alignment = EXCEL_CENTER_ALIGN
    cell.border = EXCEL_BORDER

def apply_excel_data_style(cell, align="left"):
    cell.alignment = EXCEL_CENTER_ALIGN if align == "center" else EXCEL_LEFT_ALIGN
    cell.border = EXCEL_BORDER

# --- PDF Helper Functions ---
def draw_header_footer(canvas, doc, title_text, user_name=""):
    canvas.saveState()
    
    # Draw Top Header Bar
    canvas.setFillColor(colors.HexColor('#1e3a8a'))
    canvas.rect(0, doc.pagesize[1] - 50, doc.pagesize[0], 50, fill=1, stroke=0)
    
    # Header Text
    canvas.setFillColor(colors.white)
    canvas.setFont('Helvetica-Bold', 16)
    canvas.drawString(doc.leftMargin, doc.pagesize[1] - 32, "PROMANAGE")
    
    canvas.setFont('Helvetica', 11)
    canvas.drawRightString(doc.pagesize[0] - doc.rightMargin, doc.pagesize[1] - 30, title_text)
    
    # Draw Bottom Footer Bar
    canvas.setFillColor(colors.HexColor('#f1f5f9'))
    canvas.rect(0, 0, doc.pagesize[0], 40, fill=1, stroke=0)
    
    # Footer Text
    canvas.setFillColor(colors.HexColor('#64748b'))
    canvas.setFont('Helvetica', 8)
    canvas.drawString(doc.leftMargin, 20, f"Dihasilkan pada: {datetime.datetime.now().strftime('%d %B %Y, %H:%M')}")
    if user_name:
        canvas.drawString(doc.leftMargin, 10, f"Dicetak oleh: {user_name}")
    canvas.drawRightString(doc.pagesize[0] - doc.rightMargin, 15, f"Halaman {doc.page}")
    
    # Top thin line over footer
    canvas.setStrokeColor(colors.HexColor('#cbd5e1'))
    canvas.setLineWidth(1)
    canvas.line(0, 40, doc.pagesize[0], 40)
    
    canvas.restoreState()

def get_status_color(status):
    if status.lower() == 'selesai':
        return colors.HexColor('#059669') # Emerald
    elif status.lower() == 'aktif':
        return colors.HexColor('#2563eb') # Blue
    elif status.lower() == 'tertunda':
        return colors.HexColor('#d97706') # Amber
    return colors.HexColor('#475569') # Gray


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_project_excel(request, project_id):
    """Export project data to Professional Excel format"""
    try:
        project = Proyek.objects.get(id=project_id)
    except Proyek.DoesNotExist:
        return HttpResponse("Project not found", status=404)

    wb = Workbook()
    ws = wb.active
    ws.title = "Project Report"

    # Set column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 15

    # Report Title
    ws.merge_cells('A1:F1')
    cell = ws['A1']
    cell.value = f"LAPORAN PROYEK: {project.nama.upper()}"
    cell.font = Font(bold=True, size=16, color="1E3A8A")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    
    ws.merge_cells('A2:F2')
    ws['A2'].value = f"Dihasilkan pada {datetime.datetime.now().strftime('%d %B %Y, %H:%M')} oleh {request.user.nama}"
    ws['A2'].alignment = Alignment(horizontal="center", vertical="center")
    
    # Project Details Section
    ws.append([])
    details = [
        ("Nama Proyek", project.nama, "Status", project.status),
        ("Deskripsi", project.deskripsi, "Progres", f"{hitung_progres_proyek(project)}%"),
        ("Lokasi", project.lokasi, "Supervisor", project.pengawas),
        ("Tanggal Mulai", project.tanggal_mulai.strftime('%Y-%m-%d'), "Pelaksana", project.pelaksana),
        ("Tanggal Selesai", project.tanggal_selesai.strftime('%Y-%m-%d'), "", "")
    ]
    
    for row_idx, data in enumerate(details, start=4):
        ws.cell(row=row_idx, column=1, value=data[0]).font = EXCEL_BOLD_FONT
        ws.cell(row=row_idx, column=2, value=data[1])
        ws.cell(row=row_idx, column=3, value=data[2]).font = EXCEL_BOLD_FONT
        ws.cell(row=row_idx, column=4, value=data[3])

    ws.append([])
    ws.append([])

    # Table Function
    def build_activity_table(title, activities_func, is_done):
        ws.append([title])
        title_cell = ws.cell(row=ws.max_row, column=1)
        title_cell.font = Font(bold=True, size=12, color="FFFFFF")
        title_cell.fill = PatternFill("solid", fgColor="059669" if is_done else "D97706")
        ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=6)
        
        headers = ["Pekerjaan", "Aktivitas", "Pelaksana", "Waktu", "Status"]
        ws.append(headers)
        for col_idx in range(1, 6):
            apply_excel_header_style(ws.cell(row=ws.max_row, column=col_idx))
            
        has_data = False
        for work in project.pekerjaan.all():
            acts = activities_func(work)
            for activity in acts:
                has_data = True
                row = [work.nama, activity.nama, activity.pelaksana, activity.waktu_pelaksanaan, "Selesai" if is_done else "Dalam Proses"]
                ws.append(row)
                for col_idx in range(1, 6):
                    apply_excel_data_style(ws.cell(row=ws.max_row, column=col_idx))
                    
        if not has_data:
            ws.append(["-", "Tidak ada aktivitas", "-", "-", "-"])
            for col_idx in range(1, 6):
                apply_excel_data_style(ws.cell(row=ws.max_row, column=col_idx), align="center")
        ws.append([])

    build_activity_table("AKTIVITAS SELESAI", lambda w: [a for a in w.aktivitas.all() if a.selesai], True)
    build_activity_table("AKTIVITAS DALAM PROSES", lambda w: [a for a in w.aktivitas.all() if not a.selesai], False)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Report_{project.nama.replace(" ", "_")}.xlsx"'
    wb.save(response)
    return response

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_project_pdf(request, project_id):
    """Export project data to a Professional PDF report"""
    try:
        project = Proyek.objects.get(id=project_id)
        
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="Report_{project.nama.replace(" ", "_")}.pdf"'

        doc = SimpleDocTemplate(response, pagesize=letter, rightMargin=40, leftMargin=40, topMargin=70, bottomMargin=60)
        elements = []
        styles = getSampleStyleSheet()
        
        # Define Custom Styles
        title_style = ParagraphStyle(
            'ReportTitle', parent=styles['Heading1'],
            fontSize=20, textColor=colors.HexColor('#1e3a8a'),
            spaceAfter=20, alignment=1, fontName='Helvetica-Bold'
        )
        
        section_style = ParagraphStyle(
            'SectionHeader', parent=styles['Heading2'],
            fontSize=14, textColor=colors.HexColor('#0f172a'),
            spaceBefore=15, spaceAfter=10, fontName='Helvetica-Bold',
            borderPadding=5, backColor=colors.HexColor('#f1f5f9')
        )
        
        normal_style = ParagraphStyle('CustomNormal', parent=styles['Normal'], fontSize=10, leading=14, textColor=colors.HexColor('#334155'))
        bold_style = ParagraphStyle('CustomBold', parent=styles['Normal'], fontSize=10, leading=14, fontName='Helvetica-Bold', textColor=colors.HexColor('#0f172a'))
        small_style = ParagraphStyle('CustomSmall', parent=styles['Normal'], fontSize=8, leading=11)

        # Title
        elements.append(Paragraph("LAPORAN MANAJEMEN PROYEK", title_style))
        
        # 1. Project Information Board
        elements.append(Paragraph("INFORMASI UMUM", section_style))
        
        info_data = [
            [Paragraph("<b>Nama Proyek:</b>", normal_style), Paragraph(project.nama, normal_style), Paragraph("<b>Status:</b>", normal_style), Paragraph(f"<font color='{get_status_color(project.status)}'><b>{project.status.upper()}</b></font>", normal_style)],
            [Paragraph("<b>Tanggal Mulai:</b>", normal_style), Paragraph(project.tanggal_mulai.strftime('%d %b %Y'), normal_style), Paragraph("<b>Progres:</b>", normal_style), Paragraph(f"<b>{hitung_progres_proyek(project)}%</b>", normal_style)],
            [Paragraph("<b>Tanggal Selesai:</b>", normal_style), Paragraph(project.tanggal_selesai.strftime('%d %b %Y'), normal_style), Paragraph("<b>Supervisor:</b>", normal_style), Paragraph(project.pengawas, normal_style)],
            [Paragraph("<b>Lokasi:</b>", normal_style), Paragraph(project.lokasi, normal_style), Paragraph("<b>Pelaksana:</b>", normal_style), Paragraph(project.pelaksana, normal_style)],
        ]
        info_table = Table(info_data, colWidths=[90, 175, 90, 175])
        info_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('LINEBELOW', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0'))
        ]))
        elements.append(info_table)
        
        elements.append(Spacer(1, 10))
        elements.append(Paragraph("<b>Deskripsi Proyek:</b>", normal_style))
        elements.append(Spacer(1, 5))
        elements.append(Paragraph(project.deskripsi, normal_style))
        elements.append(Spacer(1, 20))

        # 2. Detailed Works & Activities
        elements.append(Paragraph("RINCIAN PEKERJAAN & AKTIVITAS", section_style))
        works = project.pekerjaan.all()
        
        if not works:
            elements.append(Paragraph("<i>Belum ada data pekerjaan untuk proyek ini.</i>", normal_style))
        
        for idx, work in enumerate(works, 1):
            # Work Header Box
            work_header_data = [[
                Paragraph(f"<b>{idx}. {work.nama.upper()}</b>", ParagraphStyle('WHT', parent=normal_style, textColor=colors.white)),
                Paragraph(f"", ParagraphStyle('WHC', parent=normal_style, textColor=colors.white, alignment=2))
            ]]
            work_header = Table(work_header_data, colWidths=[350, 180])
            work_header.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#334155')),
                ('LEFTPADDING', (0, 0), (-1, -1), 10),
                ('RIGHTPADDING', (0, 0), (-1, -1), 10),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ]))
            elements.append(work_header)
            
            # Activities Table
            act_data = [["Aktivitas", "Waktu", "Pelaksana", "Status"]]
            activities = work.aktivitas.all()
            for act in activities:
                status_text = "<font color='#059669'><b>Selesai</b></font>" if act.selesai else "<font color='#d97706'><b>Proses</b></font>"
                act_data.append([
                    Paragraph(act.nama, small_style),
                    Paragraph(act.waktu_pelaksanaan, small_style),
                    Paragraph(act.pelaksana, small_style),
                    Paragraph(status_text, small_style)
                ])
                
                # Append Evaluation & Plan if exists
                if act.evaluasi or act.rencana_tambahan:
                    eval_plan_text = ""
                    if act.evaluasi: eval_plan_text += f"<b>Evaluasi:</b> {act.evaluasi}<br/>"
                    if act.rencana_tambahan: eval_plan_text += f"<b>Rencana:</b> {act.rencana_tambahan}"
                    act_data.append([Paragraph(eval_plan_text, small_style), "", "", ""])
            
            if not activities:
                act_data.append(["Tidak ada aktivitas", "", "", ""])
                
            act_table = Table(act_data, colWidths=[230, 100, 120, 80])
            act_table_style = [
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e2e8f0')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#0f172a')),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('TOPPADDING', (0, 0), (-1, 0), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cbd5e1')),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ]
            
            # Span rows for evaluation
            row_idx = 1
            for act in activities:
                if act.evaluasi or act.rencana_tambahan:
                    act_table_style.append(('SPAN', (0, row_idx+1), (-1, row_idx+1)))
                    act_table_style.append(('BACKGROUND', (0, row_idx+1), (-1, row_idx+1), colors.HexColor('#f8fafc')))
                    row_idx += 2
                else:
                    row_idx += 1
                    
            if not activities:
                act_table_style.append(('SPAN', (0, 1), (-1, 1)))
                act_table_style.append(('ALIGN', (0, 1), (-1, 1), 'CENTER'))

            act_table.setStyle(TableStyle(act_table_style))
            elements.append(act_table)
            elements.append(Spacer(1, 15))

        # 3. Signatures
        elements.append(Spacer(1, 30))
        elements.append(Paragraph("LEMBAR PENGESAHAN", section_style))
        
        sig_data = [
            [Paragraph("<b>Disiapkan Oleh,</b>", normal_style), Paragraph("<b>Mengetahui / Menyetujui,</b>", normal_style)],
            [Spacer(1, 60), Spacer(1, 60)],
            [Paragraph(f"<b><u>{request.user.nama}</u></b><br/>Project Manager", normal_style), 
             Paragraph(f"<b><u>{project.pengawas}</u></b><br/>Project Supervisor", normal_style)]
        ]
        sig_table = Table(sig_data, colWidths=[265, 265])
        sig_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        elements.append(sig_table)

        doc.build(elements, 
                  onFirstPage=lambda c, d: draw_header_footer(c, d, f"LAPORAN PROYEK - {project.nama[:20]}", request.user.nama),
                  onLaterPages=lambda c, d: draw_header_footer(c, d, f"LAPORAN PROYEK - {project.nama[:20]}", request.user.nama))
        return response
    except Exception as e:
        logger.error(f"Error exporting PDF: {str(e)}")
        logger.error(traceback.format_exc())
        return HttpResponse(f"Gagal membuat PDF: {str(e)}", status=500)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_all_excel(request):
    """Export all projects to Professional Excel"""
    projects = Proyek.objects.all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Global Portfolio"

    # Set column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 15

    # Report Title
    ws.merge_cells('A1:F1')
    ws['A1'].value = "LAPORAN PORTOFOLIO PROYEK GLOBAL"
    ws['A1'].font = Font(bold=True, size=16, color="1E3A8A")
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    
    ws.merge_cells('A2:F2')
    ws['A2'].value = f"Dicetak oleh: {request.user.nama} | Tanggal: {datetime.datetime.now().strftime('%d %B %Y')}"
    ws['A2'].alignment = Alignment(horizontal="center", vertical="center")
    
    ws.append([])
    
    # Summary Table
    ws.append(["RINGKASAN PROYEK"])
    ws['A4'].font = Font(bold=True, size=12)
    ws.merge_cells('A4:F4')
    
    headers = ["Nama Proyek", "Lokasi", "Mulai", "Selesai", "Progres", "Status"]
    ws.append(headers)
    for col_idx in range(1, 7):
        apply_excel_header_style(ws.cell(row=ws.max_row, column=col_idx))
        
    for p in projects:
        ws.append([p.nama, p.lokasi, p.tanggal_mulai.strftime('%Y-%m-%d'), p.tanggal_selesai.strftime('%Y-%m-%d'), f"{hitung_progres_proyek(p)}%", p.status])
        for col_idx in range(1, 7):
            apply_excel_data_style(ws.cell(row=ws.max_row, column=col_idx))
            
    ws.append([])
    ws.append([])

    # Detail Table
    ws.append(["RINCIAN PEKERJAAN PROYEK"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)
    
    for p in projects:
        ws.append([f"PROYEK: {p.nama.upper()}"])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True, color="1E3A8A")
        ws.cell(row=ws.max_row, column=1).fill = PatternFill("solid", fgColor="E2E8F0")
        ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=5)
        
        detail_headers = ["Pekerjaan", "Aktivitas", "Pelaksana", "Waktu", "Status Aktivitas"]
        ws.append(detail_headers)
        for col_idx in range(1, 6):
            cell = ws.cell(row=ws.max_row, column=col_idx)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="334155")
            cell.border = EXCEL_BORDER
            
        has_works = False
        for w in p.pekerjaan.all():
            has_works = True
            acts = w.aktivitas.all()
            if not acts:
                ws.append([w.nama, "Tidak ada aktivitas", "-", "-", "-"])
                for col_idx in range(1, 6): apply_excel_data_style(ws.cell(row=ws.max_row, column=col_idx))
            for a in acts:
                ws.append([w.nama, a.nama, a.pelaksana, a.waktu_pelaksanaan, "Selesai" if a.selesai else "Proses"])
                for col_idx in range(1, 6): apply_excel_data_style(ws.cell(row=ws.max_row, column=col_idx))
                
        if not has_works:
            ws.append(["Belum ada pekerjaan di proyek ini", "", "", "", ""])
            ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=5)
            ws.cell(row=ws.max_row, column=1).alignment = EXCEL_CENTER_ALIGN
            
        ws.append([])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Global_Portfolio_Report.xlsx"'
    wb.save(response)
    return response

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_all_pdf(request):
    """Export a comprehensive global report to Professional PDF"""
    try:
        projects = Proyek.objects.all()
        total_p = projects.count()
        done_p = projects.filter(status='Selesai').count()
        active_p = projects.filter(status='Aktif').count()
        
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="Global_Portfolio_Report.pdf"'

        doc = SimpleDocTemplate(response, pagesize=letter, rightMargin=40, leftMargin=40, topMargin=70, bottomMargin=60)
        elements = []
        styles = getSampleStyleSheet()

        title_style = ParagraphStyle(
            'ReportTitleGlobal', parent=styles['Heading1'],
            fontSize=22, textColor=colors.HexColor('#1e3a8a'),
            spaceAfter=30, alignment=1, fontName='Helvetica-Bold'
        )
        
        section_style = ParagraphStyle(
            'SectionHeaderGlobal', parent=styles['Heading2'],
            fontSize=14, textColor=colors.HexColor('#0f172a'),
            spaceBefore=20, spaceAfter=15, fontName='Helvetica-Bold',
            borderPadding=6, backColor=colors.HexColor('#f1f5f9')
        )
        
        normal_style = ParagraphStyle('CustomNormalGlobal', parent=styles['Normal'], fontSize=10, leading=14, textColor=colors.HexColor('#334155'))
        small_style = ParagraphStyle('CustomSmallGlobal', parent=styles['Normal'], fontSize=8, leading=11)

        # Title
        elements.append(Paragraph("LAPORAN PORTOFOLIO GLOBAL", title_style))
        
        # Dashboard Summary
        dash_data = [
            [Paragraph("<b>Total Proyek</b>", normal_style), Paragraph("<b>Proyek Aktif</b>", normal_style), Paragraph("<b>Proyek Selesai</b>", normal_style)],
            [Paragraph(f"<font size=18 color='#1e3a8a'><b>{total_p}</b></font>", normal_style), 
             Paragraph(f"<font size=18 color='#2563eb'><b>{active_p}</b></font>", normal_style), 
             Paragraph(f"<font size=18 color='#059669'><b>{done_p}</b></font>", normal_style)]
        ]
        dash_table = Table(dash_data, colWidths=[176, 176, 176])
        dash_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#f8fafc')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
            ('TOPPADDING', (0, 0), (-1, -1), 15),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 15),
        ]))
        elements.append(dash_table)
        elements.append(Spacer(1, 20))

        # 1. Executive Summary Table
        elements.append(Paragraph("RINGKASAN STATUS PROYEK", section_style))
        
        summary_data = [["Nama Proyek", "Lokasi", "Timeline", "Progres", "Status"]]
        for p in projects:
            summary_data.append([
                Paragraph(f"<b>{p.nama}</b>", normal_style),
                Paragraph(p.lokasi, small_style),
                Paragraph(f"{p.tanggal_mulai.strftime('%d/%m/%y')} - {p.tanggal_selesai.strftime('%d/%m/%y')}", small_style),
                Paragraph(f"<b>{hitung_progres_proyek(p)}%</b>", normal_style),
                Paragraph(f"<font color='{get_status_color(p.status)}'><b>{p.status}</b></font>", normal_style)
            ])

        st = Table(summary_data, colWidths=[180, 100, 120, 60, 70], repeatRows=1)
        st.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a8a')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cbd5e1')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
        ]))
        elements.append(st)
        elements.append(PageBreak())
        
        # 2. Detailed Breakdown
        elements.append(Paragraph("DETAIL RINCIAN SETIAP PROYEK", section_style))
        
        for p in projects:
            # Project Heading
            p_head_data = [[Paragraph(f"<b>PROYEK: {p.nama.upper()}</b>", ParagraphStyle('PH', parent=normal_style, textColor=colors.HexColor('#1e3a8a'), fontSize=12))]]
            p_head = Table(p_head_data, colWidths=[530])
            p_head.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#e0e7ff')),
                ('LEFTPADDING', (0, 0), (-1, -1), 10),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ('LINEBELOW', (0, 0), (-1, -1), 1, colors.HexColor('#818cf8')),
            ]))
            elements.append(p_head)
            elements.append(Spacer(1, 10))
            
            works = p.pekerjaan.all()
            if not works:
                elements.append(Paragraph("<i>Belum ada pekerjaan.</i>", normal_style))
                elements.append(Spacer(1, 15))
                continue
                
            for work in works:
                elements.append(Paragraph(f"<b>Pekerjaan: {work.nama}</b>", normal_style))
                elements.append(Spacer(1, 5))
                
                act_data = [["Aktivitas", "Waktu", "Pelaksana", "Status"]]
                for act in work.aktivitas.all():
                    status_str = "<font color='#059669'>Selesai</font>" if act.selesai else "<font color='#d97706'>Proses</font>"
                    act_data.append([
                        Paragraph(act.nama, small_style),
                        Paragraph(act.waktu_pelaksanaan, small_style),
                        Paragraph(act.pelaksana, small_style),
                        Paragraph(f"<b>{status_str}</b>", small_style)
                    ])
                    if act.evaluasi or act.rencana_tambahan:
                        eval_txt = ""
                        if act.evaluasi: eval_txt += f"<b>Eval:</b> {act.evaluasi}<br/>"
                        if act.rencana_tambahan: eval_txt += f"<b>Renc:</b> {act.rencana_tambahan}"
                        act_data.append([Paragraph(eval_txt, small_style), "", "", ""])
                        
                if len(act_data) == 1:
                    act_data.append(["Tidak ada aktivitas", "", "", ""])
                    
                act_table = Table(act_data, colWidths=[230, 100, 120, 80])
                act_table_style = [
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#475569')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 9),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
                    ('TOPPADDING', (0, 0), (-1, 0), 6),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cbd5e1')),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ]
                
                # Span evaluation rows
                row_idx = 1
                for act in work.aktivitas.all():
                    if act.evaluasi or act.rencana_tambahan:
                        act_table_style.append(('SPAN', (0, row_idx+1), (-1, row_idx+1)))
                        act_table_style.append(('BACKGROUND', (0, row_idx+1), (-1, row_idx+1), colors.HexColor('#f8fafc')))
                        row_idx += 2
                    else:
                        row_idx += 1
                        
                if len(act_data) == 2 and act_data[1][0] == "Tidak ada aktivitas":
                    act_table_style.append(('SPAN', (0, 1), (-1, 1)))
                    act_table_style.append(('ALIGN', (0, 1), (-1, 1), 'CENTER'))
                    
                act_table.setStyle(TableStyle(act_table_style))
                elements.append(act_table)
                elements.append(Spacer(1, 15))
            
            elements.append(Spacer(1, 20))

        # 3. Lembar Pengesahan Global
        elements.append(PageBreak())
        elements.append(Paragraph("LEMBAR PENGESAHAN PORTOFOLIO", section_style))
        elements.append(Spacer(1, 40))
        
        sig_data = [
            [Paragraph("<b>Disiapkan Oleh,</b>", normal_style), Paragraph("<b>Divalidasi Oleh,</b>", normal_style)],
            [Spacer(1, 60), Spacer(1, 60)],
            [Paragraph(f"<b><u>{request.user.nama}</u></b><br/>Project Manager", normal_style), 
             Paragraph("<b><u>_________________________</u></b><br/>Direktur / Kepala Departemen", normal_style)]
        ]
        sig_table = Table(sig_data, colWidths=[265, 265])
        sig_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        elements.append(sig_table)

        doc.build(elements, 
                  onFirstPage=lambda c, d: draw_header_footer(c, d, "LAPORAN PORTOFOLIO GLOBAL", request.user.nama),
                  onLaterPages=lambda c, d: draw_header_footer(c, d, "LAPORAN PORTOFOLIO GLOBAL", request.user.nama))
        return response
    except Exception as e:
        logger.error(f"Error exporting Global Detailed PDF: {str(e)}")
        logger.error(traceback.format_exc())
        return HttpResponse(f"Gagal membuat Laporan Global: {str(e)}", status=500)
